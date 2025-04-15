import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.copier import WorksheetCopy

# 读取 JSON 文件
with open('profile.json', 'r') as json_file:
    json_data = json.load(json_file)

# 打印 JSON 数据为单行格式
print(json.dumps(json_data, separators=(',', ':')))

# 读取 Excel 文件中名为 'Sheet1' 的工作表
excel_path = 'profile.xlsx'
df = pd.read_excel(excel_path, sheet_name='Sheet1', engine='openpyxl')

# 删除 Sheet2，如果存在
def delete_sheet_if_exists(excel_path, sheet_name='Sheet2'):
    wb = load_workbook(excel_path)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
        wb.save(excel_path)

# 删除 Sheet2
delete_sheet_if_exists(excel_path, 'Sheet2')

# 获取 JSON 中的所有配置
settings = {**json_data.get('sensorCfg', {}), **json_data.get('algCfg', {})}

# 更新 DataFrame 中的值
for index, row in df.iterrows():
    excel_key = row['E']  # 从 E 列获取值
    if pd.notna(excel_key) and excel_key in settings:  # 检查键是否存在于 JSON 数据中
        value = settings[excel_key]
        if isinstance(value, list):
            value = ' '.join(map(str, value))  # 将列表转换为字符串
        df.at[index, 'F'] = value  # 更新 E 列的值

# 使用 openpyxl 复制格式
wb = load_workbook(excel_path)
sheet1 = wb['Sheet1']

# 创建新的 Sheet2 并复制格式和数据
wb.create_sheet('Sheet2')
sheet2 = wb['Sheet2']

# 复制格式和数据
copy = WorksheetCopy(source_worksheet=sheet1, target_worksheet=sheet2)
copy.copy_worksheet()

# 将 DataFrame 数据写入到 Sheet2
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        sheet2.cell(row=r_idx, column=c_idx, value=value)

# 保存 Excel 文件
wb.save(excel_path)
